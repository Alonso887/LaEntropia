from openpyxl import Workbook, load_workbook
import datetime  
wb = None
ws1 = None

def get_name(name):#First this funtion looks for the name of the person we want data from
    name_wanted = name
    name_wanted = name_wanted.replace(" ","").lower()
    while True:
        person_cell = [7,3]#cell we re looking in and where this will start to look for people
        name_checking = ws1.cell(row=person_cell[0],column=person_cell[1]).value
        try:
            name_checking = name_checking.replace(" ","").lower()
        except:
            print("Name not founded")
            break
        if name_checking == name_wanted:
            return person_cell
            break
        else:
            person_cell[0] += 1
def get_date_list(name_cell,course_cell):#this one uses get_name to locate the persons date
    if isinstance(course_cell[0], str) and not course_cell[0].strip():
        course_cell=[5,8]#where the courses start
    date_list = []
    while True:
        data = []
        check_course = ws1.cell(row=course_cell[0],column=course_cell[1]).value#the cell we are going to check
        if check_course == None:
            break
        data.append(ws1.cell(row=name_cell[0],column=course_cell[1]).value)
        data.append(check_course)
        date_list.append(data)
        course_cell[1] += 1
    return date_list

def full_check(start_cell,name_cell,course_cell):
    if isinstance(start_cell[0], str) and not start_cell[0].strip():
        start_cell = [7,8]#Here the dates start to appear
    if isinstance(name_cell[0], str) and not name_cell[0].strip():
        name_cell = [7,3]#here the names start
    full_date_list = {}#here we save every row date with the name of the person and every date asociated with them
    while True:
        date_list = get_date_list(start_cell,course_cell)#Remember this funtion depends on get_date list
        person_list = ws1.cell(row=name_cell[0],column=name_cell[1]).value#person wich the date list corresponds to
        if person_list == None:
            break
        full_date_list[person_list] = date_list
        name_cell[0] += 1
        start_cell[0] += 1
    return full_date_list

def date_list_filter(date_list,starting_date,ending_date):#takes the date list from a person and turns it into a single list between the date ranges given
    filt_date_list = []
    for i in date_list:
        k = i[0]
        if  not isinstance(k,datetime.datetime):#only dates today baby :>  
            continue    
        if not starting_date.date() <= k.date() <= ending_date.date():
            continue
        filt_date_list.append(i)
    return filt_date_list

def full_date_list_filter(full_date_list,starting_date,ending_date):#like date list filter but for everithing, like it uses datelist filter fr 
    filtered_full_date_list = {}#it says it in the name dumbass
    for i in full_date_list:
        n = date_list_filter(full_date_list[i],starting_date,ending_date)#yes i used n as a name for a var, i'm bored as fuck
        filtered_full_date_list[i] = n
    return filtered_full_date_list

def activate_wb(file_path):#Well it sets the actual workbook, i actually feel bad for using the global
    global wb, ws1#i mean it's not bad i guess but i feel like it's a bad method or something
    wb = load_workbook(file_path)#works for this silly program but i think i should use it less
    ws1 = wb["Crew"]

def get_data(file_path,name,course_cell,start_cell,name_cell,starting_date,ending_date):
    activate_wb(file_path)    

    course_cell = course_cell.replace(" ","").split(",")
    start_cell = start_cell.replace(" ","").split(",")
    name_cell = name_cell.replace(" ","").split(",")
    
    starting_date = starting_date.replace(" ","").split("/")
    starting_date = datetime.datetime(int(starting_date[2])+2000,int(starting_date[1]),int(starting_date[0]))
    ending_date = ending_date.replace(" ","").split("/")#I added 2000 years for this to work so it reads 2023 not 23, change it if i'ts year 3000
    ending_date = datetime.datetime(int(ending_date[2])+2000,int(ending_date[1]),int(ending_date[0]))

    if isinstance(name, str) and not name.strip():
        data = full_date_list_filter(full_check(start_cell,name_cell,course_cell),starting_date,ending_date)
        return data
    else:
        data = date_list_filter(get_date_list(get_name(name),course_cell),starting_date,ending_date)
        return data
    
