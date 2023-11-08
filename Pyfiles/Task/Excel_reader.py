from openpyxl import Workbook, load_workbook

wb = load_workbook(filename=r'C:\Users\aadri\Desktop\Escuela\Training_Matrix_2006.xlsx')
ws1 = wb["Crew"]
def get_name():#First this funtion looks for the name of the person we want data from
    name_wanted = input("Wich person do you need data from?\n> ")
    name_wanted = name_wanted.replace(" ","").lower()
    while True:
        person_cell = [7,3]#cell we re looking in and where this will start to look for people
        name_checking = ws1.cell(row=person_cell[0],column=person_cell[1]).value
        try:
            name_checking = name_checking.replace(" ","").lower()
        except:
            print("Name not founded")
            break
        if name_checking == name_wanted:
            print(person_cell)
            return person_cell
            break
        else:
            person_cell[0] += 1
def get_date_list(person_cell):#this one uses get_name to locate the persons date
    course_cell=[5,8]#where the courses start
    date_list = []
    while True:
        check_course = ws1.cell(row=course_cell[0],column=course_cell[1]).value#the cell we are going to check
        if check_course == None:
            break
        date_list.append(ws1.cell(row=person_cell[0],column=course_cell[1]).value)
        course_cell[1] += 1
    return date_list

def full_check():
    start_cell = [7,8]#Here the dates start to appear
    course_cell = [5,8]#Here the courses start
    name_cell = [7,3]#here the names start
    full_date_list = {}#here we save every row date with the name of tge person and every date asociated with them
    while True:
        date_list = get_date_list(start_cell)
        person_list = ws1.cell(row=name_cell[0],column=name_cell[1]).value#person wich the date list corresponds to
        if person_list == None:
            break
        full_date_list[person_list] = date_list
        name_cell[0] += 1
        start_cell[0] += 1
    return full_date_list 


a=full_check()
print(a)
wb.save(r"C:\Users\aadri\Desktop\Escuela\Training_Matrix_2006_joke.xlsx")





